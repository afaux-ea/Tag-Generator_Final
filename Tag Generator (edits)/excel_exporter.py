from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from datetime import datetime
from utils import create_excel_styles

class ExcelExporter:
    def __init__(self):
        self.styles = create_excel_styles()

    def export_tags(self, tags, output_filename):
        wb = Workbook()
        ws = wb.active
        current_row = 1

        for tag in tags:
            current_row = self._add_tag_to_worksheet(ws, tag, current_row)
            if tag != tags[-1]:
                ws.append([''])
                current_row += 1

        self._adjust_column_widths(ws)
        wb.save(output_filename)
        return output_filename

    def _add_tag_to_worksheet(self, ws, tag, start_row):
        if 'dates' in tag:
            # Historical tag: multi-date, multi-column
            n_dates = len(tag['dates'])
            ws.append([tag['well_id']] + [""] * n_dates)
            tag_start_row = start_row
            date_headers = [
                datetime.strptime(d, "%Y-%m-%d %H:%M:%S").strftime("%B %Y") if isinstance(d, str) and len(d) >= 10 else str(d)
                for d in tag['dates']
            ]
            ws.append(["Analyte"] + date_headers)
            for analyte in tag['analytes']:
                ws.append([analyte['name']] + analyte['values'])
            self._apply_tag_formatting(ws, tag, tag_start_row)
            return start_row + len(tag['analytes']) + 2
        else:
            # Standard tag: two columns
            ws.append([tag['well_id'], ""])
            tag_start_row = start_row
            ws.append(["Analyte", tag['date']])
            for analyte in tag['analytes']:
                ws.append([analyte['name'], analyte['value']])
            self._apply_tag_formatting(ws, tag, tag_start_row)
            return start_row + len(tag['analytes']) + 2

    def _apply_tag_formatting(self, ws, tag, tag_start_row):
        if 'dates' in tag:
            n_dates = len(tag['dates'])
            end_col = n_dates + 1
            ws.merge_cells(start_row=tag_start_row, start_column=1, 
                          end_row=tag_start_row, end_column=end_col)
            well_id_cell = ws.cell(row=tag_start_row, column=1)
            well_id_cell.alignment = self.styles['center_alignment']
            well_id_cell.fill = self.styles['well_id_fill']
            analyte_row = tag_start_row + 1
            for col in range(1, end_col + 1):
                cell = ws.cell(row=analyte_row, column=col)
                cell.fill = self.styles['well_id_fill']
            for row in ws.iter_rows(min_row=tag_start_row, max_row=tag_start_row + len(tag['analytes']) + 1, min_col=1, max_col=end_col):
                for cell in row:
                    cell.border = self.styles['border']
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.row == tag_start_row or cell.column > 1:
                        cell.alignment = self.styles['center_alignment']
            for a_idx, analyte in enumerate(tag['analytes']):
                for d_idx, exceeds in enumerate(analyte.get('exceeds', [])):
                    if exceeds:
                        cell = ws.cell(row=tag_start_row + 2 + a_idx, column=2 + d_idx)
                        cell.font = self.styles['exceedance_font']
                        cell.fill = self.styles['exceedance_fill']
        else:
            # Standard (single-date) tag
            ws.merge_cells(start_row=tag_start_row, start_column=1, 
                          end_row=tag_start_row, end_column=2)
            well_id_cell = ws.cell(row=tag_start_row, column=1)
            well_id_cell.alignment = self.styles['center_alignment']
            well_id_cell.fill = self.styles['well_id_fill']
            analyte_row = tag_start_row + 1
            for col in range(1, 3):
                cell = ws.cell(row=analyte_row, column=col)
                cell.fill = self.styles['well_id_fill']
            for row in ws.iter_rows(min_row=tag_start_row, max_row=tag_start_row + len(tag['analytes']) + 1, min_col=1, max_col=2):
                for cell in row:
                    cell.border = self.styles['border']
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.row == tag_start_row or cell.column == 2:
                        cell.alignment = self.styles['center_alignment']
            for idx, analyte in enumerate(tag['analytes'], start=tag_start_row + 2):
                if analyte.get('exceeds_awqs', False):
                    cell = ws.cell(row=idx, column=2)
                    cell.font = self.styles['exceedance_font']
                    cell.fill = self.styles['exceedance_fill']

    def _adjust_column_widths(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                if not isinstance(cell, MergedCell):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            if column_letter:
                ws.column_dimensions[column_letter].width = max_length + 2