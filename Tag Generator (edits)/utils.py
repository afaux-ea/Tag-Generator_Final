from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

# Constants
DEFAULT_WINDOW_SIZE = "800x800"

# Excel styles
center_alignment = Alignment(horizontal='center', vertical='center')
border_style = Side(style='thin')
border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
exceedance_font = Font(bold=True)  # Bold black font
exceedance_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light grey
well_id_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # Light grey for Well ID

def create_excel_styles():
    return {
        'center_alignment': center_alignment,
        'border': border,
        'exceedance_font': exceedance_font,
        'exceedance_fill': exceedance_fill,
        'well_id_fill': well_id_fill
    }